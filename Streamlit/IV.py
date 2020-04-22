import tushare as ts
import pandas as pd
import openpyxl
import time
from math import log, sqrt, exp, isnan
from scipy import stats
import warnings

warnings.filterwarnings('ignore')
pd.set_option('mode.chained_assignment', None)
pro = ts.pro_api('8c5daedb3f52b5dfe11e7b2261842db0624987f22d54e8e023bbdbf2')


def get_basic_info_normal(stock_name, place):
    df_future = pro.fut_mapping(ts_code=stock_name)
    df_future_origin = pro.fut_daily(ts_code=df_future['mapping_ts_code'][0])
    print(df_future_origin)

    global mapping_name
    mapping_name = df_future_origin.ts_code[0]
    name = mapping_name[0:6]

    df_variable = pro.opt_basic(exchange=place, fields='ts_code, call_put, exercise_price')
    df_basic = df_variable[df_variable['ts_code'].str.contains(name)]
    df_basic.sort_values(by=['call_put'], ascending=False, inplace=True)
    df_basic.reset_index(drop=True, inplace=True)

    df_basic_p = df_basic[df_basic['call_put'].str.contains('P')]
    df_basic_c = df_basic[df_basic['call_put'].str.contains('C')]
    df_basic_p.reset_index(drop=True, inplace=True)
    df_basic_c.reset_index(drop=True, inplace=True)
    print(df_basic_p)
    print(df_basic_c)

    df_m = []
    i = df_basic_c.shape[0] - 1
    while i >= 0:
        df_price = pro.opt_daily(ts_code=df_basic_c.ts_code[i], trade_date='20200204')
        df_price = df_price[['close']]
        # print(df_price)
        df_m = df_price.append([df_m], ignore_index=True)

        # print(df_m)
        i = i - 1

    df_m.drop([len(df_m) - 1], inplace=True)
    print(df_m)

    global df
    df = df_basic_c[['ts_code', 'exercise_price']]
    df['close'] = df_m['close']

    df = df.dropna(axis=0, how='any')
    df.sort_values(by=['close'], ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)
    df.to_csv('d:/期货期权/隐含波动率/' + stock_name + '.csv', index=False)

    def pcr():
        sum_p_vol = 0
        sum_p_oi = 0
        for i in range(0, len(df_basic_p)):
            df_p = pro.opt_daily(ts_code=df_basic_p.ts_code[i], trade_date='20200204')
            df_vol_p = df_p[['ts_code', 'trade_date', 'exchange', 'vol']]
            df_oi_p = df_p[['ts_code', 'trade_date', 'exchange', 'oi']]
            # print(df_vol_p)
            try:
                sum_p_vol = sum_p_vol + float(df_vol_p.vol[0])
                sum_p_oi = sum_p_oi + float(df_oi_p.oi[0])
            except IndexError:
                continue
        print(sum_p_vol)
        print(sum_p_oi)

        sum_c_vol = 0
        sum_c_oi = 0
        for j in range(0, len(df_basic_c)):
            df_c = pro.opt_daily(ts_code=df_basic_c.ts_code[j], trade_date='20200204')
            df_vol_c = df_c[['ts_code', 'trade_date', 'exchange', 'vol']]
            df_oi_c = df_c[['ts_code', 'trade_date', 'exchange', 'oi']]
            # print(df_vol_c.vol[0])
            try:
                sum_c_vol = sum_c_vol + float(df_vol_c.vol[0])
                sum_c_oi = sum_c_oi + float(df_oi_c.oi[0])
            except IndexError:
                continue
        print(sum_c_vol)
        print(sum_c_oi)

        global ratio_vol, ratio_oi
        try:
            ratio_vol = sum_p_vol / sum_c_vol
            ratio_oi = sum_p_oi / sum_c_oi
        except ZeroDivisionError:
            ratio_vol = 'Nan'
            ratio_oi = 'Nan'
        print(ratio_vol)
        print(ratio_oi)

    pcr()


def get_basic_info_CZCE(stock_name, place):
    df_future = pro.fut_mapping(ts_code=stock_name)
    df_future_origin = pro.fut_daily(ts_code=df_future['mapping_ts_code'][0])
    print(df_future_origin)

    global mapping_name
    mapping_name = df_future_origin.ts_code[0]
    name = mapping_name[0:2] + mapping_name[3:6]

    df_variable = pro.opt_basic(exchange=place, fields='ts_code, call_put, exercise_price')
    df_basic = df_variable[df_variable['ts_code'].str.contains(name)]
    df_basic.sort_values(by=['call_put'], ascending=False, inplace=True)
    df_basic.reset_index(drop=True, inplace=True)

    df_basic_p = df_basic[df_basic['call_put'].str.contains('P')]
    df_basic_c = df_basic[df_basic['call_put'].str.contains('C')]
    df_basic_p.reset_index(drop=True, inplace=True)
    df_basic_c.reset_index(drop=True, inplace=True)
    print(df_basic_p)
    print(df_basic_c)

    df_m = []
    i = df_basic_c.shape[0] - 1
    while i >= 0:
        # for i in range(df_basic_c.shape[0]):
        df_price = pro.opt_daily(ts_code=df_basic_c.ts_code[i], trade_date='20200204')
        df_price = df_price[['ts_code', 'trade_date', 'close']]
        df_price = df_price[['close']]
        # print(df_price)
        df_m = df_price.append([df_m], ignore_index=True)

        # print(df_m)
        i = i - 1

    df_m.drop([len(df_m) - 1], inplace=True)
    print(df_m)

    global df
    df = df_basic_c[['ts_code', 'exercise_price']]
    df['close'] = df_m['close']

    df = df.dropna(axis=0, how='any')
    df.sort_values(by=['close'], ascending=True, inplace=True)
    df.reset_index(drop=True, inplace=True)
    df.to_csv('d:/期货期权/隐含波动率/' + stock_name + '.csv', index=False)

    def pcr():
        sum_p_vol = 0
        sum_p_oi = 0
        for i in range(0, len(df_basic_p)):
            df_p = pro.opt_daily(ts_code=df_basic_p.ts_code[i], trade_date='20200204')
            df_vol_p = df_p[['ts_code', 'trade_date', 'exchange', 'vol']]
            df_oi_p = df_p[['ts_code', 'trade_date', 'exchange', 'oi']]
            # print(df_vol_p)
            try:
                sum_p_vol = sum_p_vol + float(df_vol_p.vol[0])
                sum_p_oi = sum_p_oi + float(df_oi_p.oi[0])
            except IndexError:
                continue
        print(sum_p_vol)
        print(sum_p_oi)

        sum_c_vol = 0
        sum_c_oi = 0
        for j in range(0, len(df_basic_c)):
            df_c = pro.opt_daily(ts_code=df_basic_c.ts_code[j], trade_date='20200204')
            df_vol_c = df_c[['ts_code', 'trade_date', 'exchange', 'vol']]
            df_oi_c = df_c[['ts_code', 'trade_date', 'exchange', 'oi']]
            # print(df_vol_c.vol[0])
            try:
                sum_c_vol = sum_c_vol + float(df_vol_c.vol[0])
                sum_c_oi = sum_c_oi + float(df_oi_c.oi[0])
            except IndexError:
                continue
        print(sum_c_vol)
        print(sum_c_oi)

        global ratio_vol, ratio_oi
        try:
            ratio_vol = sum_p_vol / sum_c_vol
            ratio_oi = sum_p_oi / sum_c_oi
        except ZeroDivisionError:
            ratio_vol = 'Nan'
            ratio_oi = 'Nan'
        print(ratio_vol)
        print(ratio_oi)

    pcr()


def IV_calculation(main, filename):
    df_future = pro.fut_mapping(ts_code=main)
    df_future_origin = pro.fut_daily(ts_code=df_future['mapping_ts_code'][0], trade_date='20200204')

    price = df_future_origin.open[0]

    def bsm_call_value(s0, k, t, r, sigma):
        d1 = (log(s0 / k) + (r + 0.5 * sigma ** 2) * t) / (sigma * sqrt(t))
        d2 = (log(s0 / k) + (r - 0.5 * sigma ** 2) * t) / (sigma * sqrt(t))
        value = s0 * stats.norm.cdf(d1) - k * exp(-r * t) * stats.norm.cdf(d2)
        return value

    def bsm_vega(s0, k, t, r, sigma):
        d1 = log(s0 / k) + (r + 0.5 * sigma ** 2) * t / (sigma * sqrt(t))
        vega = s0 * stats.norm.cdf(d1, 0., 1.) * sqrt(t)

        return vega

    def bsm_call_imp_vol_newton(s0, k, t, r, c0, sigma_est, it=100):
        for i in range(it):
            sigma_est -= ((bsm_call_value(s0, k, t, r, sigma_est) - c0) / bsm_vega(s0, k, t, r, sigma_est))
        return sigma_est

    files = pd.read_csv(filename, encoding='utf-8')
    files = files.sort_values(by=['exercise_price'], ascending=1)
    files_price = files['close']
    k = files['exercise_price']

    t_10 = 10 / 252
    t_20 = 20 / 252
    t_30 = 30 / 252
    t_60 = 60 / 252

    s0 = price
    rf = 0.025

    sigma_init = 1
    global sigma_newton_10, sigma_newton_20, sigma_newton_30, sigma_newton_60
    sigma_newton_10 = []
    sigma_newton_20 = []
    sigma_newton_30 = []
    sigma_newton_60 = []

    for i in range(files.shape[0]):
        sigma_newton_10.append(bsm_call_imp_vol_newton(s0, k[i], t_10, rf, files_price[i], sigma_init))
        sigma_newton_20.append(bsm_call_imp_vol_newton(s0, k[i], t_20, rf, files_price[i], sigma_init))
        sigma_newton_30.append(bsm_call_imp_vol_newton(s0, k[i], t_30, rf, files_price[i], sigma_init))
        sigma_newton_60.append(bsm_call_imp_vol_newton(s0, k[i], t_60, rf, files_price[i], sigma_init))

    def list_deal(list_name):
        length = len(list_name)
        j = 0
        while j < length:
            a = isnan(list_name[j])
            if list_name[j] > 3 or str(a) == 'True':
                list_name[j] = (list_name[j-2] + list_name[j-3]) / 2
            j = j + 1

        sum = 0
        global avg
        avg = 0
        for x in range(0, len(list_name)):
            sum = sum + list_name[x]
            avg = sum / (i + 1)

    list_deal(sigma_newton_10)
    list_deal(sigma_newton_20)
    list_deal(sigma_newton_60)
    list_deal(sigma_newton_30)


    print('imp_vol_newton_16:')
    print(sigma_newton_10)
    print(sigma_newton_20)
    print(sigma_newton_30)
    print(sigma_newton_60)

    print(avg)


def save_config(names):
    df_all = pd.DataFrame(sigma_newton_10, columns=['10d'])
    df_all['20d'] = pd.Series(sigma_newton_20)
    df_all['30d'] = pd.Series(sigma_newton_30)
    df_all['60d'] = pd.Series(sigma_newton_60)
    df_all.apply(lambda col: col.drop_duplicates().reset_index(drop=True))

    df_all.to_csv('d:/期货期权/隐含波动率/相关数据/' + names + '.csv', index=False)


def weighted_iv(filename):
    df = pd.read_csv('d:/期货期权/隐含波动率/' + filename + '.csv')

    df_info = []
    i = len(df.ts_code) - 1

    while i >= 0:
        df_price = pro.opt_daily(ts_code=df.ts_code[i], trade_date='20200204')
        df_price = df_price[['ts_code', 'vol']]
        df_info = df_price.append([df_info], ignore_index=True)
        i = i - 1

    df_info.drop([len(df_info) - 1], inplace=True)

    sum = df_info['vol'].sum()

    df_info['ratio'] = df_info['vol'] / sum
    print(df_info)

    df_iv = pd.read_csv('d:/期货期权/隐含波动率/相关数据/' + filename + '.csv')
    #df_iv = df_iv.fillna(0)

    global weight_avg_iv
    weight_avg_iv = 0

    for j in range(0, len(df_info['ratio'])):
        weight_avg_iv = weight_avg_iv + df_iv['20d'][j] * df_info['ratio'][j]
    print(weight_avg_iv)

# 新建excel文件
file = openpyxl.Workbook()
mysheet = file.active

directionCell = mysheet.cell(row=1, column=1)
directionCell.value = 'ts_code'
directionCell = mysheet.cell(row=1, column=2)
directionCell.value = '平值隐含波动率'
directionCell = mysheet.cell(row=1, column=3)
directionCell.value = '加权隐含波动率'
directionCell = mysheet.cell(row=1, column=4)
directionCell.value = 'PCR成交量'
directionCell = mysheet.cell(row=1, column=5)
directionCell.value = 'PCR持仓量'
directionCell = mysheet.cell(row=1, column=6)
directionCell.value = '历史波动率(20)'


get_basic_info_normal('AU.SHF', 'SHFE')
IV_calculation('AU.SHF', 'd:/期货期权/隐含波动率/AU.SHF.csv')
save_config('AU.SHF')
weighted_iv('AU.SHF')
mysheet['A2'] = mapping_name
mysheet['D2'] = ratio_vol
mysheet['E2'] = ratio_oi
mysheet['B2'] = avg
mysheet['C2'] = weight_avg_iv

get_basic_info_normal('C.DCE', 'DCE')
IV_calculation('C.DCE', 'd:/期货期权/隐含波动率/C.DCE.csv')
save_config('C.DCE')
weighted_iv('C.DCE')
mysheet['A3'] = mapping_name
mysheet['D3'] = ratio_vol
mysheet['E3'] = ratio_oi
mysheet['B3'] = avg
mysheet['C3'] = weight_avg_iv

get_basic_info_CZCE('CF.ZCE', 'CZCE')
IV_calculation('CF.ZCE', 'd:/期货期权/隐含波动率/CF.ZCE.csv')
save_config('CF.ZCE')
weighted_iv('CF.ZCE')
mysheet['A4'] = mapping_name
mysheet['D4'] = ratio_vol
mysheet['E4'] = ratio_oi
mysheet['B4'] = avg
mysheet['C4'] = weight_avg_iv

get_basic_info_normal('CU.SHF', 'SHFE')
IV_calculation('CU.SHF', 'd:/期货期权/隐含波动率/CU.SHF.csv')
save_config('CU.SHF')
weighted_iv('CU.SHF')
mysheet['A5'] = mapping_name
mysheet['D5'] = ratio_vol
mysheet['E5'] = ratio_oi
mysheet['B5'] = avg
mysheet['C5'] = weight_avg_iv

time.sleep(60)
get_basic_info_normal('I.DCE', 'DCE')
IV_calculation('I.DCE', 'd:/期货期权/隐含波动率/I.DCE.csv')
save_config('I.DCE')
weighted_iv('I.DCE')
mysheet['A6'] = mapping_name
mysheet['D6'] = ratio_vol
mysheet['E6'] = ratio_oi
mysheet['B6'] = avg
mysheet['C6'] = weight_avg_iv

#time.sleep(60)
get_basic_info_normal('M.DCE', 'DCE')
IV_calculation('M.DCE', 'd:/期货期权/隐含波动率/M.DCE.csv')
save_config('M.DCE')
weighted_iv('M.DCE')
mysheet['A7'] = mapping_name
mysheet['D7'] = ratio_vol
mysheet['E7'] = ratio_oi
mysheet['B7'] = avg
mysheet['C7'] = weight_avg_iv

get_basic_info_CZCE('MA.ZCE', 'CZCE')
IV_calculation('MA.ZCE', 'd:/期货期权/隐含波动率/MA.ZCE.csv')
save_config('MA.ZCE')
weighted_iv('MA.ZCE')
mysheet['A8'] = mapping_name
mysheet['D8'] = ratio_vol
mysheet['E8'] = ratio_oi
mysheet['B8'] = avg
mysheet['C8'] = weight_avg_iv

get_basic_info_CZCE('RM.ZCE', 'CZCE')
IV_calculation('RM.ZCE', 'd:/期货期权/隐含波动率/RM.ZCE.csv')
save_config('RM.ZCE')
weighted_iv('RM.ZCE')
mysheet['A9'] = mapping_name
mysheet['D9'] = ratio_vol
mysheet['E9'] = ratio_oi
mysheet['B9'] = avg
mysheet['C9'] = weight_avg_iv

time.sleep(60)
get_basic_info_normal('RU.SHF', 'SHFE')
IV_calculation('RU.SHF', 'd:/期货期权/隐含波动率/RU.SHF.csv')
save_config('RU.SHF')
weighted_iv('RU.SHF')
mysheet['A10'] = mapping_name
mysheet['D10'] = ratio_vol
mysheet['E10'] = ratio_oi
mysheet['B10'] = avg
mysheet['C10'] = weight_avg_iv

get_basic_info_CZCE('SR.ZCE', 'CZCE')
IV_calculation('SR.ZCE', 'd:/期货期权/隐含波动率/SR.ZCE.csv')
save_config('SR.ZCE')
weighted_iv('SR.ZCE')
mysheet['A11'] = mapping_name
mysheet['D11'] = ratio_vol
mysheet['E11'] = ratio_oi
mysheet['B11'] = avg
mysheet['C11'] = weight_avg_iv

time.sleep(60)
get_basic_info_CZCE('TA.ZCE', 'CZCE')
IV_calculation('TA.ZCE', 'd:/期货期权/隐含波动率/TA.ZCE.csv')
save_config('TA.ZCE')
weighted_iv('TA.ZCE')
mysheet['A12'] = mapping_name
mysheet['D12'] = ratio_vol
mysheet['E12'] = ratio_oi
mysheet['B12'] = avg
mysheet['C12'] = weight_avg_iv

file.save('./test.xlsx')


df_iv = pd.read_excel('./test.xlsx')
df_hv20 = pd.read_excel('d:/期货期权/处理数据/All.xlsx')

df_iv['历史波动率(20)'] = df_hv20['HV20_AVG']
df_iv.to_excel('./test.xlsx')

df_file = pd.read_excel('./test.xlsx', index_col=0)
df_file.to_excel('d:/期货期权/隐含波动率/总表.xlsx', index=False)
