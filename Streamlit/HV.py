import shutil
import os
import tushare as ts
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from win32com.client import Dispatch
from pandas.plotting import register_matplotlib_converters

register_matplotlib_converters()


def just_open(filename):
    xlApp = Dispatch('Excel.Application')
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


def get_future_daily(future_name):
    pro = ts.pro_api('8c5daedb3f52b5dfe11e7b2261842db0624987f22d54e8e023bbdbf2')
    df_future = pro.fut_mapping(ts_code=future_name)
    df = pro.fut_daily(ts_code=df_future['mapping_ts_code'][0], start_date='20190101', end_date='')

    df.sort_values(by=['trade_date'], ascending=True, inplace=True)
    df.to_excel('d:/期货期权/原始数据/'+df_future['mapping_ts_code'][0]+'2019行情.xlsx', index=False, columns=['ts_code', 'trade_date', 'open', 'high', 'low', 'close'])

    df = df[['ts_code', 'trade_date', 'open', 'high', 'low', 'close']]
    df.index = range(1, len(df) + 1)

    df['NextClose'] = df['close'].shift(-1)
    df['ReturnRate'] = df['close'].pct_change(1)
    df = df[['ts_code', 'trade_date', 'close', 'NextClose', 'ReturnRate']]

    df.to_excel('d:/期货期权/处理数据/回报率/'+df_future['mapping_ts_code'][0]+'2019回报率数据.xlsx', encoding='utf_8_sig')
    print(df)

    filename = df_future['mapping_ts_code'][0]

    def hv_calculate():
        file = openpyxl.load_workbook('d:/期货期权/处理数据/回报率/'+filename+'2019回报率数据.xlsx')
        mysheet = file.active

        row = mysheet.max_row
        print(row)

        directionCell = mysheet.cell(row=1, column=7)
        directionCell.value = "HV5"
        a = 3
        while a <= row - 4:
            mysheet['G' + str(a + 4)] = '=STDEV(F' + str(a) + ':F' + str(a + 4) + ') * SQRT(252)'
            a = a + 1


        directionCell = mysheet.cell(row=1, column=9)
        directionCell.value = "HV10"
        b = 3
        while b <= row - 9:
            mysheet['I' + str(b + 9)] = '=STDEV(F' + str(b) + ':F' + str(b + 9) + ') * SQRT(252)'
            b = b + 1


        directionCell = mysheet.cell(row=1, column=11)
        directionCell.value = "HV15"
        c = 3
        while c <= row - 14:
            mysheet['K' + str(c + 14)] = '=STDEV(F' + str(c) + ':F' + str(c + 14) + ') * SQRT(252)'
            c = c + 1


        directionCell = mysheet.cell(row=1, column=13)
        directionCell.value = "HV20"
        d = 3
        while d <= row - 19:
            mysheet['M' + str(d + 19)] = '=STDEV(F' + str(d) + ':F' + str(d + 19) + ') * SQRT(252)'
            d = d + 1

        directionCell = mysheet.cell(row=1, column=15)
        directionCell.value = "HV30"
        e = 3
        while e <= row - 29:
            mysheet['O' + str(e + 29)] = '=STDEV(F' + str(e) + ':F' + str(e + 29) + ') * SQRT(252)'
            e = e + 1

        directionCell = mysheet.cell(row=1, column=17)
        directionCell.value = "HV60"
        f = 3
        while f <= row - 59:
            mysheet['Q' + str(f + 59)] = '=STDEV(F' + str(f) + ':F' + str(f + 59) + ') * SQRT(252)'
            f = f + 1

        directionCell = mysheet.cell(row=1, column=19)
        directionCell.value = "HV90"
        g = 3
        while g <= row - 89:
            mysheet['S' + str(g + 89)] = '=STDEV(F' + str(g) + ':F' + str(g + 89) + ') * SQRT(252)'
            g = g + 1

        directionCell = mysheet.cell(row=1, column=21)
        directionCell.value = "HV120"
        h = 3
        while h <= row - 119:
            mysheet['U' + str(h + 119)] = '=STDEV(F' + str(h) + ':F' + str(h + 119) + ') * SQRT(252)'
            h = h + 1

        file.save('d:/期货期权/处理数据/波动率/' + filename + '2019历史波动率数据.xlsx')

    def hv_awl():
        file = openpyxl.load_workbook('d:/期货期权/处理数据/波动率/' + filename + '2019历史波动率数据.xlsx')
        mysheet = file.active

        directionCell = mysheet.cell(row=1, column=23)
        directionCell.value = "HV"

        mysheet['W2'] = 'HV005'
        mysheet['W3'] = 'HV010'
        mysheet['W4'] = 'HV015'
        mysheet['W5'] = 'HV020'
        mysheet['W6'] = 'HV030'
        mysheet['W7'] = 'HV060'
        mysheet['W8'] = 'HV090'
        mysheet['W9'] = 'HV120'

        directionCell = mysheet.cell(row=1, column=24)
        directionCell.value = "10分位"
        mysheet['X2'] = '=PERCENTILE(G:G,0.1)'
        mysheet['X3'] = '=PERCENTILE(I:I,0.1)'
        mysheet['X4'] = '=PERCENTILE(K:K,0.1)'
        mysheet['X5'] = '=PERCENTILE(M:M,0.1)'
        mysheet['X6'] = '=PERCENTILE(O:O,0.1)'
        mysheet['X7'] = '=PERCENTILE(Q:Q,0.1)'
        mysheet['X8'] = '=PERCENTILE(S:S,0.1)'
        mysheet['X9'] = '=PERCENTILE(U:U,0.1)'

        directionCell = mysheet.cell(row=1, column=25)
        directionCell.value = "25分位"
        mysheet['Y2'] = '=PERCENTILE(G:G,0.25)'
        mysheet['Y3'] = '=PERCENTILE(I:I,0.25)'
        mysheet['Y4'] = '=PERCENTILE(K:K,0.25)'
        mysheet['Y5'] = '=PERCENTILE(M:M,0.25)'
        mysheet['Y6'] = '=PERCENTILE(O:O,0.25)'
        mysheet['Y7'] = '=PERCENTILE(Q:Q,0.25)'
        mysheet['Y8'] = '=PERCENTILE(S:S,0.25)'
        mysheet['Y9'] = '=PERCENTILE(U:U,0.25)'

        directionCell = mysheet.cell(row=1, column=26)
        directionCell.value = "50分位"
        mysheet['Z2'] = '=PERCENTILE(G:G,0.5)'
        mysheet['Z3'] = '=PERCENTILE(I:I,0.5)'
        mysheet['Z4'] = '=PERCENTILE(K:K,0.5)'
        mysheet['Z5'] = '=PERCENTILE(M:M,0.5)'
        mysheet['Z6'] = '=PERCENTILE(O:O,0.5)'
        mysheet['Z7'] = '=PERCENTILE(Q:Q,0.5)'
        mysheet['Z8'] = '=PERCENTILE(S:S,0.5)'
        mysheet['Z9'] = '=PERCENTILE(U:U,0.5)'

        directionCell = mysheet.cell(row=1, column=27)
        directionCell.value = "75分位"
        mysheet['AA2'] = '=PERCENTILE(G:G,0.75)'
        mysheet['AA3'] = '=PERCENTILE(I:I,0.75)'
        mysheet['AA4'] = '=PERCENTILE(K:K,0.75)'
        mysheet['AA5'] = '=PERCENTILE(M:M,0.75)'
        mysheet['AA6'] = '=PERCENTILE(O:O,0.75)'
        mysheet['AA7'] = '=PERCENTILE(Q:Q,0.75)'
        mysheet['AA8'] = '=PERCENTILE(S:S,0.75)'
        mysheet['AA9'] = '=PERCENTILE(U:U,0.75)'

        directionCell = mysheet.cell(row=1, column=28)
        directionCell.value = "90分位"
        mysheet['AB2'] = '=PERCENTILE(G:G,0.9)'
        mysheet['AB3'] = '=PERCENTILE(I:I,0.9)'
        mysheet['AB4'] = '=PERCENTILE(K:K,0.9)'
        mysheet['AB5'] = '=PERCENTILE(M:M,0.9)'
        mysheet['AB6'] = '=PERCENTILE(O:O,0.9)'
        mysheet['AB7'] = '=PERCENTILE(Q:Q,0.9)'
        mysheet['AB8'] = '=PERCENTILE(S:S,0.9)'
        mysheet['AB9'] = '=PERCENTILE(U:U,0.9)'

        directionCell = mysheet.cell(row=1, column=14)
        directionCell.value = "HV20_AVG"
        mysheet['N2'] = '=AVERAGE(M:M)'
        directionCell = mysheet.cell(row=1, column=16)
        directionCell.value = "HV30_AVG"
        mysheet['P2'] = '=AVERAGE(O:O)'
        directionCell = mysheet.cell(row=1, column=18)
        directionCell.value = "HV60_AVG"
        mysheet['R2'] = '=AVERAGE(Q:Q)'
        directionCell = mysheet.cell(row=1, column=20)
        directionCell.value = "HV90_AVG"
        mysheet['T2'] = '=AVERAGE(S:S)'

        file.save('d:/期货期权/处理数据/波动率/' + filename + '2019历史波动率数据.xlsx')

    def hv_pic():
        df_hv = pd.read_excel('d:/期货期权/处理数据/波动率/' + filename + '2019历史波动率数据.xlsx')
        df_hv['trade_date'] = pd.to_datetime(df_hv['trade_date'], format='%Y%m%d')
        plt.figure(figsize=(14, 7))

        plt.plot(df_hv['trade_date'], df_hv['HV10'], linewidth=1.2, label='HV10')
        plt.plot(df_hv['trade_date'], df_hv['HV20'], linewidth=1.2, label='HV20')
        plt.plot(df_hv['trade_date'], df_hv['HV30'], linewidth=1.2, label='HV30')

        plt.xlabel('date', size=12)
        plt.ylabel('%', size=15)

        plt.title(filename + ' HV', size=12)
        plt.legend(loc='upper left')
        plt.savefig('d:/期货期权/图/K线图/' + filename + 'HV.png')
        plt.close()

    def awl_pic():
        df_awl = pd.read_excel('d:/期货期权/处理数据/波动率/' + filename + '2019历史波动率数据.xlsx')
        plt.figure(figsize=(14, 7))

        plt.plot(df_awl['HV'][:8], df_awl['10分位'][:8], linewidth=1.2, label='10%')
        plt.plot(df_awl['HV'][:8], df_awl['25分位'][:8], linewidth=1.2, label='25%')
        plt.plot(df_awl['HV'][:8], df_awl['50分位'][:8], linewidth=1.2, label='50%')
        plt.plot(df_awl['HV'][:8], df_awl['75分位'][:8], linewidth=1.2, label='75%')
        plt.plot(df_awl['HV'][:8], df_awl['90分位'][:8], linewidth=1.2, label='90%')

        plt.title(filename + ' HV', size=12)
        plt.legend(loc='upper left')
        plt.savefig('d:/期货期权/图/锥图/' + filename + '锥.png')
        plt.close()

    hv_calculate()
    hv_awl()

    just_open('d:/期货期权/处理数据/波动率/' + filename + '2019历史波动率数据.xlsx')

    hv_pic()
    awl_pic()


print('请关闭所有文件夹。')
judge = input('请确保您的文件已备份，备份完成请输入Y，未完成请输入N:')
if judge == 'Y' or 'y':
    if os.path.exists('d:/期货期权'):
        shutil.rmtree('d:/期货期权')  # 删除非空子目录

        os.makedirs(r'd:/期货期权/原始数据')
        os.makedirs(r'd:/期货期权/隐含波动率')
        os.makedirs(r'd:/期货期权/图')
        os.makedirs(r'd:/期货期权/hiv')

        os.makedirs(r'd:/期货期权/处理数据/回报率')
        os.makedirs(r'd:/期货期权/处理数据/波动率')

        os.makedirs(r'd:/期货期权/隐含波动率/相关数据')
        os.makedirs(r'd:/期货期权/图/K线图')
        os.makedirs(r'd:/期货期权/图/锥图')

    get_future_daily('CU.SHF')
    get_future_daily('AU.SHF')
    get_future_daily('RU.SHF')

    get_future_daily('M.DCE')
    get_future_daily('I.DCE')
    get_future_daily('C.DCE')

    get_future_daily('TA.ZCE')
    get_future_daily('SR.ZCE')
    get_future_daily('RM.ZCE')
    get_future_daily('MA.ZCE')
    get_future_daily('CF.ZCE')

else:
    print('请备份完成后运行!')


dirs = 'd:/期货期权/处理数据/波动率'
files = os.listdir(dirs)

df1 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[0])
df1 = df1[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df1 = df1.iloc[:1]

df2 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[1])
df2 = df2[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df2 = df2.iloc[:1]

df3 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[2])
df3 = df3[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df3 = df3.iloc[:1]

df4 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[3])
df4 = df4[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df4 = df4.iloc[:1]

df5 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[4])
df5 = df5[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df5 = df5.iloc[:1]

df6 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[5])
df6 = df6[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df6 = df6.iloc[:1]

df7 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[6])
df7 = df7[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df7 = df7.iloc[:1]

df8 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[7])
df8 = df8[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df8 = df8.iloc[:1]

df9 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[8])
df9 = df9[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df9 = df9.iloc[:1]

df10 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[9])
df10 = df10[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df10 = df10.iloc[:1]

df11 = pd.read_excel('d:/期货期权/处理数据/波动率/' + files[10])
df11 = df11[['ts_code', 'HV20_AVG', 'HV30_AVG', 'HV60_AVG', 'HV90_AVG']]
df11 = df11.iloc[:1]

df = df1.append([df2, df3, df4, df5, df6, df7, df8, df9, df10, df11], ignore_index=True)
print(df)
df.to_excel('d:/期货期权/处理数据/All.xlsx', index=False)
